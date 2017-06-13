from openpyxl import load_workbook
from openpyxl import Workbook

INPUT_FILE = 'google_form_pure.xlsx'
OUTPUT_FILE = 'output.xlsx'


class GetParams:
    """Return names of params (margin, shape, etc)"""
    def __init__(self, row):
        self.row = [str(cell.value).replace('\n', '').strip() for cell in row]

        # BiRad params

        self.mass_shape = self.get_params(10, 13)
        self.lymph_nodes = self.get_params(44, 46)

        # Mammography params
        self.mammo_mass_margin = self.get_params(13, 18)
        self.mammo_mass_density = self.get_params(18, 22)
        self.mammo_calcifications_typically_benign = self.get_params(22, 31)
        self.mammo_calcifications_suspicious_morphology = self.get_params(31, 35)
        self.mammo_calcifications_suspicious_distribution = self.get_params(35, 40)
        self.mammo_asymmetries = self.get_params(40, 44)

        # US params
        self.us_mass_margin = self.get_params(57, 63)
        self.us_mass_echo_pattern = self.get_params(63, 69)
        self.us_mass_posterior_features = self.get_params(69, 73)
        self.us_calcifications = self.get_params(73, 76)
        self.us_special_cases = self.get_params(88, 97)

        # MRI params
        self.mri_mass_margin = self.get_params(102, 106)
        self.mri_mass_internal_enhacement = self.get_params(106, 110)
        self.mri_features = self.get_params(110, 113)
        self.mri_kinetic_curve_assessment = self.get_params(113, 119)
        self.mri_nme_distribution = self.get_params(120, 126)
        # Non-mass enhancement (NME) - Internal enhancement patterns
        self.mri_nme_iep = self.get_params(126, 130)
        # Non-enhancing findings
        self.mri_nef = self.get_params(130, 137)
        # Fat containing lesions
        self.mri_fcl = self.get_params(139, 144)

    def get_params(self, start, stop):
        return [x.split('[')[1][:-1] for x in self.row[start:stop]]


class SetParams:
    """Create list of params for output file"""
    def __init__(self, params_name, row):
        self.row = row
        x = 10
        y = 13
        b = 18
        c = 22
        d = 31
        e = 35
        f = 40
        g = 44
        h = 46

        i = 54
        j = 57
        j2 = 63
        l = 69
        m = 73
        n = 76
        o = 78
        p = 88
        q = 97


        r = 99
        s = 102
        t = 106
        u = 110
        v = 113
        w = 119
        z = 120
        z1 = 126
        z2 = 130
        z3 = 137
        z4 = 139
        z5 = 144


        probabilities = ['Typical', 'Possible', 'Atypical', 'None', 'Pathogenomonic', 'Unrelated', 'Negative', 'Ignore']
        short_notation = ['t', 'p', 'a', 'none', 'pat', 'u', 'n', 'i']

        # Mammography params
        self.mammo_params = {
            'Mass': {
                'Shape': {
                    probabilities[k]: self.return_params(x, y, short_notation[k], params_name.mass_shape)
                    for k in range(len(probabilities))
                    },
                'Margin': {
                    probabilities[k]: self.return_params(y, b, short_notation[k], params_name.mammo_mass_margin)
                    for k in range(len(probabilities))
                    },
                'Density': {
                    probabilities[k]: self.return_params(b, c, short_notation[k], params_name.mammo_mass_density)
                    for k in range(len(probabilities))
                }
            },
            'Calcifications': {
                'Typically benign': {
                    probabilities[k]: self.return_params(c, d, short_notation[k],
                    params_name.mammo_calcifications_typically_benign)
                    for k in range(len(probabilities))
                },
                'Suspicious morphology': {
                    probabilities[k]: self.return_params(d, e, short_notation[k],
                    params_name.mammo_calcifications_suspicious_morphology)
                    for k in range(len(probabilities))
                },
                'Distribution': {
                    probabilities[k]: self.return_params(e, f, short_notation[k],
                    params_name.mammo_calcifications_suspicious_distribution)
                    for k in range(len(probabilities))
                }
            },
            'Assymetry': {
                'Assymetry': {
                    probabilities[k]: self.return_params(f, g, short_notation[k], params_name.mammo_asymmetries)
                    for k in range(len(probabilities))
                }
            },
            'Lymph nodes': {
                'Lymph nodes': {
                    probabilities[k]: self.return_params(g, h, short_notation[k], params_name.lymph_nodes)
                    for k in range(len(probabilities))
                }
            },
        }

        # US params
        self.us_params = {
            'Mass': {
                'Shape': {
                    probabilities[k]: self.return_params(i, j, short_notation[k], params_name.mass_shape)
                    for k in range(len(probabilities))
                },
                'Margin': {
                    probabilities[k]: self.return_params(j, j2, short_notation[k], params_name.us_mass_margin)
                    for k in range(len(probabilities))
                },
                'Echo': { #63, 69
                    probabilities[k]: self.return_params(j2, l, short_notation[k], params_name.us_mass_echo_pattern)
                    for k in range(len(probabilities))
                },
                'Posterior': { #69, 73
                    probabilities[k]: self.return_params(l, m, short_notation[k], params_name.us_mass_posterior_features)
                    for k in range(len(probabilities))
                }
            },
            'Calcifications US': {
                'Calcifications US': { #73, 76
                    probabilities[k]: self.return_params(m, n, short_notation[k], params_name.us_calcifications)
                    for k in range(len(probabilities))
                }
            },

            'Lymph nodes': {
                'Lymph nodes': { #76, 78
                    probabilities[k]: self.return_params(n, o, short_notation[k], params_name.lymph_nodes)
                    for k in range(len(probabilities))
                }
            },
            'Special cases': {
                'Special cases': { #88, 97
                    probabilities[k]: self.return_params(p, q, short_notation[k], params_name.us_special_cases)
                    for k in range(len(probabilities))
                }
            }
        }


        # MRI params
        self.mri_params = {
            'Mass': {
                'Shape': { #99, 102
                    probabilities[k]: self.return_params(r, s, short_notation[k], params_name.mass_shape)
                    for k in range(len(probabilities))
                },
                'Margin': { #102, 106
                    probabilities[k]: self.return_params(s, t, short_notation[k], params_name.mri_mass_margin)
                    for k in range(len(probabilities))
                },
                'Internal enhacement': { #106, 110
                    probabilities[k]: self.return_params(t, u, short_notation[k], params_name.mri_mass_internal_enhacement)
                    for k in range(len(probabilities))
                }
            },
            'MRI features': {
                'MRI features': { #110, 113
                    probabilities[k]: self.return_params(u, v, short_notation[k], params_name.mri_features)
                    for k in range(len(probabilities))
                }
            },
            'Kinetic curve assessment': {
                'Kinetic curve assessment': { #113, 119
                    probabilities[k]: self.return_params(v, w, short_notation[k], params_name.mri_kinetic_curve_assessment)
                    for k in range(len(probabilities))
                }
            },
            'Non-mass enhancement (NME)': {
                'Distribution': { #120, 126
                    probabilities[k]: self.return_params(z, z1, short_notation[k], params_name.mri_nme_distribution)
                    for k in range(len(probabilities))
                },
                'Internal enhancement patterns': { #126, 130
                    probabilities[k]: self.return_params(z1, z2, short_notation[k], params_name.mri_nme_iep)
                    for k in range(len(probabilities))
                }
            },
            'Non-enhancing findings': {
                'Non-enhancing findings': { #130, 137
                    probabilities[k]: self.return_params(z2, z3, short_notation[k], params_name.mri_nef)
                    for k in range(len(probabilities))
                }
            },
            'Lymph nodes': {
                'Lymph nodes': { #137, 139
                    probabilities[k]: self.return_params(z3, z4, short_notation[k], params_name.lymph_nodes)
                    for k in range(len(probabilities))
                }
            },
            'Fat containing lesions': {
                'Fat containing lesions': { #139, 144
                    probabilities[k]: self.return_params(z4, z5, short_notation[k], params_name.mri_fcl)
                    for k in range(len(probabilities))
                }
            },
        }

    def return_params(self, start, stop, params_type, param_name):
        """
        :param params_type: t - typical, p - possible, a - atypical,
         pat - Pathogenomonic, u - Unrelated, n - Negative, i - Ignore
        :return: string
        """
        pt = {
            't': 'Typical', 'p': 'Possible', 'a': 'Atypical', 'none': 'None',
            'pat': 'Pathogenomonic', 'u': 'Unrelated',
            'n': 'Negative', 'i': 'Ignore'
        }
        string = ', '.join([param_name[count]
                            for count, cell in enumerate(self.row[start:stop])
                            if pt[params_type] in cell])
        return string


def read_file():
    """Return data from input file as dictionary"""
    # Read file
    wb = load_workbook(INPUT_FILE)
    ws = wb.worksheets[0]
    return ws


def ch_none(string):
    if 'None' not in string:
        return True
    else:
        return False


def get_dictionary(file_data):
    """Return data from file as list of dictionaries"""
    data_iter_rows = list(file_data.iter_rows())
    params = GetParams(data_iter_rows[0])
    data_list = []
    for rows in data_iter_rows[1:]:
        row = [str(cell.value).replace('\n', '').strip() for cell in rows]
        rel_modalities = [x.strip() for x in row[4].split(',') if ch_none(x)]
        unique_finding = ', '.join([x for x in row[163:173] if ch_none(x)])
        params_list = SetParams(params, row)
        dict_birad = {'Typical': 50, 'Possible': 30, 'None':1, 'Ignore': 'Ignore'}
        d = {'Name': row[1], 'Condition description': row[2],
             'Relevant modalities': rel_modalities,
             'Unique findings': unique_finding,
             'mammo_params': params_list.mammo_params,
             'us_params': params_list.us_params,
             'mri_params': params_list.mri_params,
             'birad[0]': dict_birad[row[153]],
             'birad[1]': dict_birad[row[154]],
             'birad[2]': dict_birad[row[155]],
             'birad[3]': dict_birad[row[156]],
             'birad[4]': dict_birad[row[157]],
             'birad[5]': dict_birad[row[158]],
             'birad[6]': dict_birad[row[159]],
             'Associated conditions': row[160],
             'Differential diagnosis': row[162],
             }
        data_list.append(d)
    return data_list


def get_row_nr(file_row, param_name):
    """Return dict of params for output file"""
    if param_name == 'mammo_params':
        name = 'Mammography'
    elif param_name == 'us_params':
        name = 'US'
    else:
        name = 'MRI'
    data = []
    for mp_key, mv in file_row[param_name].items():
        for sub_mp_key, _ in file_row[param_name][mp_key].items():
            d1 = {'Relevant findings': mp_key, 'Parameters': sub_mp_key,
                  'Relevant modalities': name}
            for smk, smv in file_row[param_name][mp_key][sub_mp_key].items():
                if smv:
                    d2 = {smk: smv}
                    d1 = {**d1, **d2}
            data.append(d1)
    return data


def get_output_list(file_row):
    """Return list of dict for output file"""
    data = []
    list_mammo = get_row_nr(file_row, 'mammo_params')
    list_us = get_row_nr(file_row, 'us_params')
    list_mri = get_row_nr(file_row, 'mri_params')
    all_params = list_mammo + list_us + list_mri
    for x in range(len(all_params)):
        d = {'Name': '', 'Condition description': '',
             'Unique findings': '','Additional info':'','birad[0]':'','birad[1]':'',
             'birad[2]':'','birad[3]':'','birad[4]':'','birad[5]':'','birad[6]':'', 'Associated conditions': '', 'Differential diagnosis': ''}
        nd = {**d, **all_params[x]}
        if x == 0:
            nd['Name'] = file_row['Name']
            nd['Condition description'] = file_row['Condition description']
            nd['Unique findings'] = file_row['Unique findings']
            nd['birad[0]'] = file_row['birad[0]']
            nd['birad[1]'] = file_row['birad[1]']
            nd['birad[2]'] = file_row['birad[2]']
            nd['birad[3]'] = file_row['birad[3]']
            nd['birad[4]'] = file_row['birad[4]']
            nd['birad[5]'] = file_row['birad[5]']
            nd['birad[6]'] = file_row['birad[6]']
            nd['Associated conditions'] = file_row['Associated conditions']
            nd['Differential diagnosis'] = file_row['Differential diagnosis']
        data.append(nd)
    return data


def save_output(output_list):
    """Save data in output file"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Sheet1'
    # Create title for columns
    columns_titles = ['Name', 'Condition description', 'birad[0]','birad[1]','birad[2]','birad[3]','birad[4]','birad[5]','birad[6]','Relevant modalities',
                      'Relevant findings', 'Unique findings','Additional info',
                      'Parameters', 50, 30, 1, 'General',
                      'Pathogenomonic', 'Unrelated', 'Negative',
                      'Ignore', 'Associated conditions', 'Differential diagnosis', 'Notes']
    ws1.append(columns_titles)
    # Create list for output file
    for ol in output_list:
        for o in ol:
            cr_list = create_list(o)
            ws1.append(cr_list)
    wb.save(filename=OUTPUT_FILE)


def create_list(row):
    """Return list for output file"""
    name = row['Name']
    cd = row['Condition description']
    br0 = row['birad[0]']
    br1 = row['birad[1]']
    br2 = row['birad[2]']
    br3 = row['birad[3]']
    br4 = row['birad[4]']
    br5 = row['birad[5]']
    br6 = row['birad[6]']
    rm = row['Relevant modalities']
    rf = row['Relevant findings']
    uf = row['Unique findings']
    ai = row['Additional info']
    params = row['Parameters']
    try:
        t = row['Typical']
    except:
        t = ''
    try:
        p = row['Possible']
    except:
        p = ''
    try:
        a = row['Atypical']
    except:
        a = ''
    try:
        none = row['None']
    except:
        none = ''
    try:
        pat = row['Pathogenomonic']
    except:
        pat = ''
    try:
        u = row['Unrelated']
    except:
        u = ''
    try:
        n = row['Negative']
    except:
        n = ''
    try:
        i = row['Ignore']
    except:
        i = ''
    try:
        notes = row['Notes']
    except:
        notes = ''
    ac = row['Associated conditions']
    dd = row['Differential diagnosis']
    return [name, cd, br0, br1, br2, br3, br4, br5, br6, rm, rf, uf, ai, params, t, p, a, none, pat, u, n, i, ac, dd, notes]


def main():
    file_data = read_file()
    dictionary_data = get_dictionary(file_data)
    output_list = [get_output_list(x) for x in dictionary_data]
    save_output(output_list)


if __name__ == '__main__':
    main()
