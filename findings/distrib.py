import pandas as pd
import numpy as np
import random
import json
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np
import math

# define values check and append to arr
# define probability array
# read excel

df = pd.read_excel("output.xlsx")
wb = load_workbook('output.xlsx')
ws = wb.get_sheet_by_name('Sheet1')  # Define worksheet


def get_dic_from_two_lists(keys, values):
    return {keys[i]: values[i] for i in range(len(keys))}


# Define function to normalize arr values
def normalize(items):
    problist = [x / sum(items) for x in items]


# def probslist

def concatvals(row, start, stop):
    prob_head = list(df)[start:stop]
    width = stop - start
    col = start
    val_arr = []
    prob_arr = []
    for i in range(width):
        value_temp = df.iloc[row - 2, col]
        if isinstance(value_temp, float) is False:
            value = [x.strip() for x in value_temp.split(',')]
            len_val = len(value)
            prob_arr += [prob_head[i] for _ in range(len_val)]
            val_arr += value[0:len_val]
        col += 1
    randparameter = random.choices(val_arr, prob_arr, k=1)
    return randparameter

def grab_data(r, s, x, y):
    ps = [concatvals(r+s, x, y)]
    return ps

def create_rep(arr, dict1, row_data, condname, modality):
    params = []
    to_json = []
    if condname == 'Mass' and modality == 'Mammography':
        for i in range(len(arr)):
            params += grab_data(row_data, 0, 14, 19)
            row_data += 1

    elif condname == 'Calcifications' and modality == 'Mammography':
        for i in range(len(arr)):
            params += grab_data(row_data, 3, 14, 19)
            row_data += 1

    elif condname == 'Assymetry' and modality == 'Mammography':
        for i in range(len(arr)):
            params += grab_data(row_data, 6, 14, 19)
            row_data += 1

    elif condname == 'Lymph nodes' and modality == 'Mammography':
        for i in range(len(arr)):
            params += [concatvals(row_data + 7, 14, 19)]
            row_data += 1

    elif condname == 'Mass' and modality == 'US':
        for i in range(len(arr)):
            params += [concatvals(row_data + 8, 14, 19)]
            row_data += 1

    elif condname == 'Calcifications US' and modality == 'US':
        for i in range(len(arr)):
            params += [concatvals(row_data + 12, 14, 19)]
            row_data += 1

    elif condname == 'Lymph nodes' and modality == 'US':
        for i in range(len(arr)):
            params += [concatvals(row_data + 13, 14, 19)]
            row_data += 1

    elif condname == 'Special cases' and modality == 'US':
        for i in range(len(arr)):
            params += [concatvals(row_data + 14, 14, 19)]
            row_data += 1

    elif condname == 'Mass' and modality == 'MRI':
        for i in range(len(arr)):
            params += [concatvals(row_data + 15, 14, 19)]
            row_data += 1

    elif condname == 'MRI features' and modality == 'MRI':
        for i in range(len(arr)):
            params += [concatvals(row_data + 18, 14, 19)]
            row_data += 1

    elif condname == 'Kinetic curve assessment' and modality == 'MRI':
        for i in range(len(arr)):
            params += [concatvals(row_data + 19, 14, 19)]
            row_data += 1

    elif condname == 'Non-mass enhancement (NME)' and modality == 'MRI':
        for i in range(len(arr)):
            params += [concatvals(row_data + 20, 14, 19)]
            row_data += 1

    elif condname == 'Non-enhancing findings' and modality == 'MRI':
        for i in range(len(arr)):
            params += [concatvals(row_data + 22, 14, 19)]
            row_data += 1

    elif condname == 'Lymph nodes' and modality == 'MRI':
        for i in range(len(arr)):
            params += [concatvals(row_data + 22, 14, 19)]
            row_data += 1

    elif condname == 'Fat containing lesions' and modality == 'MRI':
        for i in range(len(arr)):
            params += [concatvals(row_data + 23, 14, 19)]
            row_data += 1

    data = get_dic_from_two_lists(arr, params)
    dict1.update(data)
    to_json += [dict1]
    return to_json


def get_name(infile):
    with open(infile, 'r') as f:
        contents_of_file = f.read()
        lines = contents_of_file.splitlines()
    line_number = random.randrange(0, len(lines))
    person_name = lines[line_number]
    return person_name

def get_numcond():
    names = len(df.Name.unique())
    return names


def get_cond_name():
    name_arr = df.Name.unique()
    n = list(name_arr)
    n_arr = []
    for i in range(len(name_arr)):
        if (isinstance(n[i], float)) is False:
            n_arr += [n[i]]
    rand_cond_name = random.choice(n_arr)
    return rand_cond_name


def check_row(cond_name):
    from xlrd import open_workbook
    book = open_workbook("output.xlsx")
    for sheet in book.sheets():
        for rowidx in range(sheet.nrows):
            row = sheet.row(rowidx)
            for colidx, cell in enumerate(row):
                if cell.value == cond_name:
                    print("condition name is: ", cond_name)
                    return rowidx + 1


# Create random with parameter of report numbers
def generate_report(infile):
    # for i in range(items):
    a = np.array([[i.value for i in j] for j in ws['C1':'I1']]).ravel()
    b = np.array([[i.value for i in j] for j in ws['C2':'I2']]).ravel()
    # Read BiRads Probabilities into list
    # Read BiRads into list
    person_name = get_name(infile)
    p_id = random.randrange(100)
    p_age = random.randrange(25, 65)
    br_p = normalize(b)
    br = random.choices(a, br_p, k=1)
    name = get_cond_name()
    names = get_numcond()
    row = check_row(name)
    "create list of values and slice empty entities from list"
    rm = df['Relevant modalities'].values.tolist()[0:26]
    # r = 'Mammography'
    r = random.choice(rm)
    dict_report = {'Id': p_id, 'First name': person_name, 'Age': p_age, 'Condition Name': name, 'BiRad': br,
                   'Relevant Modality': r}


    # mammo params
    if r == 'Mammography':
        f_list = df['Relevant findings'].values.tolist()[0:8]
        f = random.choice(f_list)
        dict_report.update({'Relevant finding': f})
        iter_params_mass = ['Shape', 'Margin', 'Density']
        iter_params_calc = ['Typically benign', 'Suspicious morphology', 'Distribution']
        iter_params_a = ['Assymetry']
        iter_params_lymph = ['Lymph nodes']
        if f == 'Mass':
             report = create_rep(iter_params_mass, dict_report, row, f, r)
        elif f == 'Calcifications':
            report = create_rep(iter_params_calc, dict_report, row, f, r)
        elif f == 'Assymetry':
            report = create_rep(iter_params_a, dict_report, row, f, r)
        else:
            report = create_rep(iter_params_lymph, dict_report, row, f, r)
    elif r == 'US':
        f_list = df['Relevant findings'].values.tolist()[8:15]
        f = random.choice(f_list)
        dict_report.update({'Relevant finding': f})
        us_params_mass = ['Shape', 'Margin', 'Echo', 'Posterior']
        us_params_calc = ['Calcifications']
        us_params_l_nodes = ['Lymph Nodes']
        us_params_sp_cases = ['Special Cases']
        if f == 'Mass':
            report = create_rep(us_params_mass, dict_report, row, f, r)
        elif f == 'Calcifications US':
            report = create_rep(us_params_calc, dict_report, row, f, r)
        elif f == 'Lymph nodes':
            report = create_rep(us_params_l_nodes, dict_report, row, f, r)
        else:
            report = create_rep(us_params_sp_cases, dict_report, row, f, r)
    elif r == 'MRI':
        f_list = df['Relevant findings'].values.tolist()[15:25]
        mri_params_mass = ['Shape', 'Margin', 'Internal enhancement']
        mri_params_mri_f = ['MRI features']
        mri_params_kin_c_a = ['Kinetic curve assessment']
        mri_params_nme = ['Distribution', 'Internal enhancement patterns']
        mri_params_nef = ['Non-enhancing patterns']
        mri_params_l_nodes = ['Lymph Nodes']
        mri_params_fcl = ['Fat containing lesions']
        f = random.choice(f_list)
        dict_report.update({'Relevant finding': f})
        if f == 'Mass':
            report = create_rep(mri_params_mass, dict_report, row, f, r)
        elif f == 'MRI features':
            report = create_rep(mri_params_mri_f, dict_report, row, f, r)
        elif f == 'Kinetic curve assessment':
            report = create_rep(mri_params_kin_c_a, dict_report, row, f, r)
        elif f == 'Non-mass enhancement (NME)':
            report = create_rep(mri_params_nme, dict_report, row, f, r)
        elif f == 'Non-enhancing findings':
            report = create_rep(mri_params_nef, dict_report, row, f, r)
        elif f == 'Lymph nodes':
            report = create_rep(mri_params_l_nodes, dict_report, row, f, r)
        else:
            report = create_rep(mri_params_fcl, dict_report, row, f, r)
    print(report)


def main():
    for i in range(1):
        generate_report("first-names.txt")


main()
