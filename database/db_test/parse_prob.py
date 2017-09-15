from openpyxl import load_workbook
from openpyxl import Workbook
import xlsxwriter
import itertools
import pandas as pd

INPUT_FILE = 'output.xlsx'


df = pd.read_excel("output.xlsx")
wb = load_workbook(filename=INPUT_FILE)
ws = wb['cond']

workbook = xlsxwriter.Workbook('data.xlsx')
worksheet = workbook.add_worksheet()


def get_data(start, stop, col):
    mylist = [ws.cell(row=i,column= col).value for i in range(start,stop)]
    findings = [v for i,v in enumerate(mylist) if mylist.index(v) == i]
    return findings

def get_data_row(start, stop, row):
    mylist = [ws.cell(row = row, column = i).value for i in range(start, stop)]
    values = [v for i,v in enumerate(mylist) if mylist.index(v) == i]
    values = filter(None, values)
    values = [item for word in values for item in word.split(',')]
    return values

def write_from_dict(row, col, dict):
    row = row
    col = col
    for key in dict.keys():
        worksheet.write(row, col, key)
        for item in dict[key]:
            worksheet.write(row, col + 1, item)
            row += 1



if __name__ == '__main__':

    mammo_findings_list = get_data(2,10, 11)
    us_findings_list = get_data(10, 17, 11)
    mri_findings_list = get_data(17, 27, 11)

    mammo_params_list = [get_data(2, 5, 14), get_data(5,8,14),
                         get_data(8, 9, 14), get_data(9,10, 14)]
    #print(mammo_params_list[0][1], mammo_params_list[1], mammo_params_list[2], mammo_params_list[3])
    mammo_data = dict(zip(mammo_findings_list,mammo_params_list))

    us_params_list = [get_data(10, 14, 14), get_data(14, 15, 14),
                      get_data(15, 16, 14), get_data(16, 17, 14)]
    us_data = dict(zip(us_findings_list,us_params_list))

    mri_params_list = [get_data(17, 20, 14), get_data(20, 21, 14),
                       get_data(21, 22, 14), get_data(22, 24, 14),
                       get_data(24, 25, 14), get_data(25, 26, 14),
                       get_data(26, 27, 14)]
    mri_data = dict(zip(mri_findings_list, mri_params_list))

    #mammo = write_from_dict(1, 1, mammo_data)
    #us = write_from_dict(9, 1, us_data)
    #mri = write_from_dict(16, 1, mri_data)


    shape_val = get_data_row(15, 20, 2)
    margin_val = get_data_row(15, 20, 3)
    density_val = get_data_row(15, 20, 4)

    calc_t_b_val = get_data_row(15, 20, 5)
    calc_s_m_val = get_data_row(15, 20, 6)
    calc_d_bal = get_data_row(15, 20, 7)

    print(len(margin_val))
    print(calc_t_b_val)

    shape_dict = dict(zip(mammo_params_list[0],[shape_val]))
    margin_dict = dict(zip(([mammo_params_list[0][1]]),[margin_val]))
    density_dict = dict(zip(([mammo_params_list[0][2]]),[density_val]))

    calc_t_b_dict = dict(zip(([mammo_params_list[1][0]]),[calc_t_b_val]))

    mass_shape = write_from_dict(1, 2, shape_dict)
    mass_margin =  write_from_dict((len(shape_val)+1), 2, margin_dict)
    mass_density = write_from_dict((len(margin_val)+1)+len(shape_val), 2, density_dict)

    calc_t_b = write_from_dict((len(margin_val)+1)+len(shape_val)+len(density_val),
                               2, calc_t_b_dict)



    workbook.close()

